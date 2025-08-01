import os
import time
import openpyxl
from dotenv import load_dotenv
import google.generativeai as genai
import cohere
import evaluate
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
# === Load API keys ===
load_dotenv()
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
co = cohere.Client(os.getenv("COHERE_API_KEY"))

# === Prompts to test ===
PROMPTS = [
    "How does the struct '_lprint_job_s' work in lprint-job.h?",
    "What happens when register_device is called in device.c?",
    "Explain the flow of main() function in printer-main.c",
    "What is the role of lprintCreateJob in lprint-job.c?",
    "How are printers registered in lprint-system.c?",
    "Explain what lprintMarkDirty does in lprint-job.c",
]

# === Metric
rouge = evaluate.load("rouge")

# === Excel setup
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Model Comparison"
sheet.append([
    "Prompt",
    "Gemini Answer",
    "Gemini Time (s)",
    "Cohere Answer",
    "Cohere Time (s)",
    "ROUGE-L (F1)",
    "Better Model"
])

# === Run comparison
for prompt in PROMPTS:
    print(f"\n🚀 Prompt: {prompt}")

    # === Gemini
    gemini_start = time.time()
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        gemini_response = model.generate_content(prompt).text.strip()
    except Exception as e:
        gemini_response = f"❌ Gemini Error: {e}"
    gemini_time = time.time() - gemini_start
    print(f"⏱️ Gemini took {gemini_time:.2f}s")

    # === Cohere
    cohere_start = time.time()
    try:
        cohere_response = co.generate(
            model="command-r-plus",
            prompt=prompt,
            max_tokens=1024,
            temperature=0.3
        ).generations[0].text.strip()
    except Exception as e:
        cohere_response = f"❌ Cohere Error: {e}"
    cohere_time = time.time() - cohere_start
    print(f"⏱️ Cohere took {cohere_time:.2f}s")

    # === Metric (ROUGE-L)
    try:
        rouge_score = rouge.compute(
            predictions=[gemini_response],
            references=[cohere_response]
        )
        rouge_l = round(rouge_score["rougeL"], 4)
    except:
        rouge_l = "❌"

    # === Automatically decide better model
    if isinstance(rouge_l, float):
        if rouge_l >= 0.85:
            better_model = "Both"
        elif rouge_score["rougeL"] > 0.5:
            better_model = "Gemini ≈ Cohere"
        elif "Error" in cohere_response:
            better_model = "Gemini"
        elif "Error" in gemini_response:
            better_model = "Cohere"
        else:
            better_model = "Gemini" if rouge_score["rougeL"] > 0.3 else "Cohere"
    else:
        better_model = "Undetermined"

    # === Save to Excel
    sheet.append([
        prompt,
        gemini_response,
        round(gemini_time, 2),
        cohere_response,
        round(cohere_time, 2),
        rouge_l,
        better_model
    ])
current_row = sheet.max_row
for col in range(1, sheet.max_column + 1):
    cell = sheet.cell(row=current_row, column=col)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

for col in range(1, sheet.max_column + 1):
    col_letter = get_column_letter(col)
    sheet.column_dimensions[col_letter].width = 50
# === Save workbook
wb.save("model_comparison_results.xlsx")
print("\n✅ Results saved to model_comparison_results.xlsx")
